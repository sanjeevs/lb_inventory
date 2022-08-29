import re
import os
import openpyxl

WILCO_HEADER_COLUMN = ['SKU', 'LOCATION', 'Batch #', 'Description', 'OnHand', 'Allocated', 'Available']

def extract_wilco_header_row_index(sheet):
    """Extract the header row of the table by matching the cell value SKU.
       Assumes that SKU is present in one of the cells of the header row.
       Becase of merged cell the row can have None in some of the cells.

    :param sheet: active sheet

    :rtype: row index 1 based
    """

    header_row_idx = None

    for row in sheet.iter_rows(max_row=10):
        col_values = [cell.value for cell in row]
        num_matches = 0
        for v in col_values:
            if v in WILCO_HEADER_COLUMN:
                num_matches += 1
            if num_matches == len(WILCO_HEADER_COLUMN):
                header_row_idx = row[0].row
                break
        if header_row_idx:
            break

    if not header_row_idx:
        raise ValueError("Could not find wilco column header in report")

    return header_row_idx

def create_inventory_database(sheet, header_row_idx):
    """Extract all the information into a database.
       The column indices are hardwired to 0, 6, 9

    """

    inventory = []
    type_item = ""
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if 'Details' in row[0]:
            continue
            
        if row[0] == "":
            # Some rows are for totals
            continue
        inventory.append([row[0].replace('\n',''), row[6], row[9]])

    if len(inventory) == 0:
        raise ValueError("Wilco inventory database has no entries")    

    return inventory

def save_inventory_database_to_xlsx(fname, db):
    """Save inventory database to a xlsx file."""

    outbook = openpyxl.Workbook()
    outsheet = outbook.active
    outsheet["A1"] = "SKU"
    outsheet["B1"] = "Description"
    outsheet["C1"] = "OnHand"

    for row, (sku, description, onhand) in enumerate(db, start=2):
        outsheet[f"A{row}"] = sku
        outsheet[f"A{row}"].number_format = '@'
        outsheet[f"B{row}"] = description
        outsheet[f"C{row}"] = onhand
    outbook.save(fname)

def create_categories_database(sheet):
    """Create a assoc array of categories from a sheet."""

    categories = {}
    for row in range(2, sheet.max_row + 1):
        sku = sheet['A' + str(row)].value
        item_type = sheet['C' + str(row)].value
        item_category = sheet['D' + str(row)].value
        
        categories[sku] = (item_type, item_category)

    if len(categories) == 0:
        raise ValueError("No entry read in categories")
    else:
        print(f"Found {len(categories)} matching categories")

    return categories