import re
import os
import openpyxl
from pathlib import Path


def extract_pricing_header_row_index(sheet):
    """Extract the header row of the table by matching the cell value SKU.
       Assumes that SKU is present in one of the cells of the header row.

    :param sheet: active sheet

    :rtype: row index 1 based
    """

    header_row_idx = None

    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        if "SKU" in values or "SKU NO" in values:
            header_row_idx = row[0].row           # 1-based
            break
    if header_row_idx:
        print(f"Found header row at {header_row_idx}")
    else:
        raise ValueError("Could not find header row for the table")
    return header_row_idx

def extract_pricing_header_col_indices(sheet, row_idx):
    """Extract the column indices for all the header fields"""

    col_indices = {}

    header_row = sheet[row_idx]
    values = [cell.value for cell in header_row]
    print(values)
    col_indices['SKU'] = values.index('SKU') + 1                # 1-based
    col_indices['Description'] = values.index("DESCRIPTION OF GOODS") + 1
    col_indices['wt_pc_in_gms'] = values.index("wt/pcs in gms") + 1
    col_indices['pcs_per_case'] = values.index("pcs/case") + 1
    col_indices['price_per_pc'] = values.index('price/pcs') + 1
    return col_indices

def create_pricing_database(sheet, header_row_idx, header_col_indices):
    """Extract all the information into a database"""

    db = []
    header_fields = header_col_indices.keys()

    # Start reading the sheet after the header row.
    for row in sheet.iter_rows(min_row=header_row_idx + 1):

        # Must have a value in first cell of the row
        if row[0].value:
            entry = {}
            for f in header_fields:
                entry[f] = row[header_col_indices[f] -1].value

            db.append(entry)

    if len(db):
        print(f"Found {len(db)} entries in the database")
    else:
        raise ValueError("Found no entries in database")
    
    return db

def save_pricing_database_to_xlsx(fname, db):
    """Save the pricing database to a xlsx fname."""

    outbook = openpyxl.Workbook()
    outsheet = outbook.active

    header_fields = db[0].keys()
    for idx, header in enumerate(header_fields):
        outsheet.cell(row=1, column=idx + 1).value = header

    for idx, entry in enumerate(db):
        col_idx = 1
        outsheet.cell(row=2 + idx, column=1).number_format = '@'
        for f in header_fields:
            outsheet.cell(row=2 + idx, column=col_idx).value = entry[f]
            col_idx += 1

    print(f"Saving output file '{fname}'")
    outbook.save(fname)
